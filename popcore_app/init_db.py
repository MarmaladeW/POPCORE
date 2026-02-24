"""
init_db.py - Parse both Excel files and initialize the SQLite database.
Run this once (or whenever the Excel files are updated) to refresh the DB.
"""
import openpyxl
import sqlite3
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'popcore.db')

MASTER_FILE = os.path.join(BASE_DIR, 'copy of 11.xlsx')
DETAIL_FILE = os.path.join(BASE_DIR, 'POP_CORE_v3.xlsx')

# Sheets to skip in the detail file
SKIP_SHEETS = {'📋 总览 Index'}


def parse_master(path):
    """Parse copy of 11.xlsx → dict keyed by SKU with 记账名 and full name."""
    if not os.path.exists(path):
        print(f'ERROR: Master file not found: {path}')
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = {}
    for row in ws.iter_rows(values_only=True):
        sku = row[0]
        if sku and str(sku).strip().startswith('SP'):
            sku = str(sku).strip()
            records[sku] = {
                'jizhanming': str(row[1]).strip() if row[1] else '',
                'full_name_master': str(row[2]).strip() if row[2] else '',
                'brand_master': str(row[3]).strip() if row[3] else '',
            }
    return records


def parse_detail(path):
    """Parse POP_CORE_v3.xlsx → list of product dicts."""
    if not os.path.exists(path):
        print(f'ERROR: Detail file not found: {path}')
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    products = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Row 0 = title, Row 1 = headers, Row 2+ = data
        for row in rows[2:]:
            if not any(cell for cell in row):
                continue
            # Columns: 名称, 价格, 发售时间, 发售体数/限量, 发售渠道, 隐藏款, 款式特点, 产品类型, 记账名, SKU
            name_cn_en = str(row[0]).strip() if row[0] else ''
            price = row[1]
            release_date = str(row[2]).strip() if row[2] else ''
            edition_size = str(row[3]).strip() if row[3] else ''
            channel = str(row[4]).strip() if row[4] else ''
            hidden = str(row[5]).strip() if row[5] else ''
            style_notes = str(row[6]).strip() if row[6] else ''
            product_type = str(row[7]).strip() if row[7] else ''
            jizhanming_detail = str(row[8]).strip() if row[8] else ''
            sku = str(row[9]).strip() if row[9] else ''

            if not name_cn_en and not sku:
                continue

            try:
                price_val = float(price) if price is not None else None
            except (TypeError, ValueError):
                price_val = None

            products.append({
                'ip_series': sheet_name,
                'name_cn_en': name_cn_en,
                'price': price_val,
                'release_date': release_date,
                'edition_size': edition_size,
                'channel': channel,
                'hidden': hidden,
                'style_notes': style_notes,
                'product_type': product_type,
                'jizhanming_detail': jizhanming_detail,
                'sku': sku,
            })
    return products


def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute('PRAGMA journal_mode=WAL')
    cur = con.cursor()

    cur.executescript('''
        CREATE TABLE IF NOT EXISTS products (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sku         TEXT UNIQUE,
            name_cn_en  TEXT,
            jizhanming  TEXT,
            price       REAL,
            ip_series   TEXT,
            product_type TEXT,
            brand       TEXT,
            release_date TEXT,
            edition_size TEXT,
            channel     TEXT,
            hidden      TEXT,
            style_notes TEXT,
            notes       TEXT DEFAULT '',
            search_blob TEXT   -- lowercase concat for fuzzy search
        );

        CREATE TABLE IF NOT EXISTS inventory (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id  INTEGER NOT NULL REFERENCES products(id),
            date        TEXT NOT NULL,          -- YYYY-MM-DD
            opening_qty INTEGER NOT NULL DEFAULT 0,
            restock_qty INTEGER NOT NULL DEFAULT 0,
            sold_qty    INTEGER NOT NULL DEFAULT 0,
            closing_qty INTEGER,               -- actual physical count
            expected_qty INTEGER,              -- opening + restock - sold
            discrepancy INTEGER,               -- closing - expected
            notes       TEXT DEFAULT '',
            UNIQUE(product_id, date)
        );

        CREATE INDEX IF NOT EXISTS idx_products_sku ON products(sku);
        CREATE INDEX IF NOT EXISTS idx_inventory_date ON inventory(date);
    ''')

    master = parse_master(MASTER_FILE)
    detail_list = parse_detail(DETAIL_FILE)

    # Build merged product records
    # Start with detail file as primary source
    seen_skus = set()
    all_products = []

    for p in detail_list:
        sku = p['sku']
        m = master.get(sku, {})

        # Prefer jizhanming from master (copy of 11), fall back to detail
        jzm = m.get('jizhanming') or p['jizhanming_detail']
        brand = m.get('brand_master') or p['ip_series']
        name = p['name_cn_en'] or m.get('full_name_master', '')

        search_blob = ' '.join([
            (sku or '').lower(),
            (jzm or '').lower(),
            (name or '').lower(),
            (brand or '').lower(),
            (p['product_type'] or '').lower(),
            (p['ip_series'] or '').lower(),
        ])

        all_products.append({
            'sku': sku,
            'name_cn_en': name,
            'jizhanming': jzm,
            'price': p['price'],
            'ip_series': p['ip_series'],
            'product_type': p['product_type'],
            'brand': brand,
            'release_date': p['release_date'],
            'edition_size': p['edition_size'],
            'channel': p['channel'],
            'hidden': p['hidden'],
            'style_notes': p['style_notes'],
            'notes': '',
            'search_blob': search_blob,
        })
        seen_skus.add(sku)

    # Add any products only in master file (not in detail)
    for sku, m in master.items():
        if sku not in seen_skus:
            jzm = m.get('jizhanming', '')
            name = m.get('full_name_master', '')
            brand = m.get('brand_master', '')
            search_blob = ' '.join([sku.lower(), jzm.lower(), name.lower(), brand.lower()])
            all_products.append({
                'sku': sku,
                'name_cn_en': name,
                'jizhanming': jzm,
                'price': None,
                'ip_series': '',
                'product_type': '',
                'brand': brand,
                'release_date': '',
                'edition_size': '',
                'channel': '',
                'hidden': '',
                'style_notes': '',
                'notes': '',
                'search_blob': search_blob,
            })

    cur.executemany('''
        INSERT INTO products (sku, name_cn_en, jizhanming, price, ip_series, product_type,
                              brand, release_date, edition_size, channel, hidden,
                              style_notes, notes, search_blob)
        VALUES (:sku, :name_cn_en, :jizhanming, :price, :ip_series, :product_type,
                :brand, :release_date, :edition_size, :channel, :hidden,
                :style_notes, :notes, :search_blob)
        ON CONFLICT(sku) DO UPDATE SET
            name_cn_en  = excluded.name_cn_en,
            jizhanming  = excluded.jizhanming,
            price       = excluded.price,
            ip_series   = excluded.ip_series,
            product_type = excluded.product_type,
            brand       = excluded.brand,
            release_date = excluded.release_date,
            edition_size = excluded.edition_size,
            channel     = excluded.channel,
            hidden      = excluded.hidden,
            style_notes = excluded.style_notes,
            search_blob = excluded.search_blob
    ''', all_products)

    con.commit()
    con.close()
    print(f'DB initialized at {DB_PATH}')
    print(f'Products loaded: {len(all_products)}')


if __name__ == '__main__':
    init_db()
