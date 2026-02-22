"""
export_excel.py - Export the full POPCORE product database to a formatted Excel file.
Run from the popcore_app directory: python export_excel.py
"""
import sqlite3
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'popcore.db')
EXPORT_DIR = os.path.dirname(BASE_DIR)  # one level up, in POPCORE folder

timestamp = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT_FILE = os.path.join(EXPORT_DIR, f'POPCORE_export_{timestamp}.xlsx')


def export():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # Fetch all products with stock info (LEFT JOIN so products without stock still appear)
    cur.execute('''
        SELECT
            p.sku,
            p.jizhanming,
            p.name_cn_en,
            p.brand,
            p.ip_series,
            p.product_type,
            p.price,
            p.release_date,
            p.edition_size,
            p.channel,
            p.hidden,
            p.style_notes,
            p.boxes_per_dan,
            COALESCE(s.upstairs_dan, 0)  AS upstairs_dan,
            COALESCE(s.instore_dan, 0)   AS instore_dan,
            (COALESCE(s.upstairs_dan, 0) + COALESCE(s.instore_dan, 0)) AS total_dan,
            p.notes
        FROM products p
        LEFT JOIN stock s ON s.product_id = p.id
        ORDER BY p.ip_series, p.sku
    ''')
    rows = cur.fetchall()
    con.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'POPCORE库存总表'

    # ── Styles ──────────────────────────────────────────────
    header_font   = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    header_fill   = PatternFill('solid', fgColor='1A1A2E')   # dark navy
    accent_fill   = PatternFill('solid', fgColor='E8F4FD')   # light blue for even rows
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin          = Side(style='thin', color='CCCCCC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    price_fmt     = '#,##0.00'
    int_fmt       = '#,##0'

    # ── Headers ──────────────────────────────────────────────
    headers = [
        ('SKU',         10),
        ('记账名',       18),
        ('产品名称',     30),
        ('品牌',         14),
        ('IP系列',       16),
        ('产品类型',     12),
        ('价格(¥)',      10),
        ('发售时间',     12),
        ('发售体数',     10),
        ('发售渠道',     14),
        ('隐藏款',        8),
        ('款式特点',     20),
        ('每端盒数',      8),
        ('楼上(端)',      8),
        ('店内(端)',      8),
        ('合计(端)',      8),
        ('备注',         20),
    ]

    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    # ── Data rows ─────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        fill = accent_fill if row_idx % 2 == 0 else None

        values = [
            row['sku'],
            row['jizhanming'],
            row['name_cn_en'],
            row['brand'],
            row['ip_series'],
            row['product_type'],
            row['price'],
            row['release_date'],
            row['edition_size'],
            row['channel'],
            row['hidden'],
            row['style_notes'],
            row['boxes_per_dan'],
            row['upstairs_dan'],
            row['instore_dan'],
            row['total_dan'],
            row['notes'],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = left_align if col_idx in (2, 3, 5, 10, 12, 17) else center_align
            if fill:
                cell.fill = fill

            # Number formatting
            if col_idx == 7 and value is not None:   # price
                cell.number_format = price_fmt
            elif col_idx in (13, 14, 15, 16) and value is not None:  # qty columns
                cell.number_format = int_fmt

        ws.row_dimensions[row_idx].height = 18

    # ── Auto-filter ───────────────────────────────────────────
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows) + 1}'

    # ── Summary sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet('汇总统计')
    ws2['A1'] = '导出时间'
    ws2['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws2['A2'] = '产品总数'
    ws2['B2'] = len(rows)

    # Count products with stock
    with_stock = sum(1 for r in rows if (r['upstairs_dan'] or 0) + (r['instore_dan'] or 0) > 0)
    ws2['A3'] = '有库存品种'
    ws2['B3'] = with_stock
    ws2['A4'] = '楼上总端数'
    ws2['B4'] = sum((r['upstairs_dan'] or 0) for r in rows)
    ws2['A5'] = '店内总端数'
    ws2['B5'] = sum((r['instore_dan'] or 0) for r in rows)
    ws2['A6'] = '全部总端数'
    ws2['B6'] = sum((r['total_dan'] or 0) for r in rows)

    for r in range(1, 7):
        ws2.cell(row=r, column=1).font = Font(bold=True)
        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 20

    wb.save(OUTPUT_FILE)
    print(f'✓ 导出完成: {OUTPUT_FILE}')
    print(f'  产品总数: {len(rows)}')
    print(f'  有库存品种: {with_stock}')


if __name__ == '__main__':
    export()
